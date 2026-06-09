import pandas as pd
from pathlib import Path
from openpyxl import load_workbook
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side


# =========================
# CONFIG
# =========================

INPUT_FILE = "Agreement.xlsx"
OUTPUT_FILE = "valid_triples_analysis.xlsx"

ID_COL = "ID"
SENTENCE_COL = "sentence"

MODELS = [
    {
        "name": "GPT4omini",
        "triple_col": "GPT4ominiresults",
        "status_col": "Status",
    },
    {
        "name": "LLaMa3",
        "triple_col": "LLaMa3results",
        "status_col": "Status.1",
    },
    {
        "name": "GPT55",
        "triple_col": "GPT55results",
        "status_col": "Status.2",
    },
    {
        "name": "DeepSeekR1",
        "triple_col": "DeepSeekR1results",
        "status_col": "Status.3",
    },
]


# =========================
# UTILS
# =========================

def is_real_triple(value):
    """
    True se la cella contiene una vera tripla.
    Esclude NaN, stringhe vuote e [].
    """
    if pd.isna(value):
        return False

    value = str(value).strip()

    if value == "":
        return False

    if value == "[]":
        return False

    return True


def normalize_status(value):
    if pd.isna(value):
        return ""
    return str(value).strip().upper()


# =========================
# LOAD DATA
# =========================

df = pd.read_excel(INPUT_FILE)

# Propaga ID e sentence sulle righe vuote appartenenti alla stessa frase
df[ID_COL] = df[ID_COL].ffill()
df[SENTENCE_COL] = df[SENTENCE_COL].ffill()

# Rende ID intero se possibile
df[ID_COL] = df[ID_COL].astype(int)


# =========================
# SHEET 1 — VALID_MARKED_TRIPLES
# =========================

valid_marked_rows = []

for sentence_id, group in df.groupby(ID_COL, sort=False):
    sentence = group[SENTENCE_COL].iloc[0]

    valid_triples_by_model = {}

    max_len = 0

    for model in MODELS:
        model_name = model["name"]
        triple_col = model["triple_col"]
        status_col = model["status_col"]

        valid_triples = group.loc[
            group[status_col].apply(normalize_status).eq("VALID")
            & group[triple_col].apply(is_real_triple),
            triple_col
        ].astype(str).tolist()

        valid_triples_by_model[model_name] = valid_triples
        max_len = max(max_len, len(valid_triples))

    # Se nessun modello ha triple VALID, manteniamo comunque una riga per la frase
    if max_len == 0:
        max_len = 1

    for i in range(max_len):
        row = {
            "id_frase": sentence_id if i == 0 else "",
            "frase": sentence if i == 0 else "",
        }

        for model in MODELS:
            model_name = model["name"]
            triples = valid_triples_by_model[model_name]

            row[f"{model_name}_valid_triples"] = triples[i] if i < len(triples) else ""

        valid_marked_rows.append(row)

valid_marked_df = pd.DataFrame(valid_marked_rows)


# =========================
# SHEET 2 — VALID_COUNT
# =========================

valid_count_rows = []

for sentence_id, group in df.groupby(ID_COL, sort=False):
    sentence = group[SENTENCE_COL].iloc[0]

    row = {
        "id_frase": sentence_id,
        "frase": sentence,
    }

    for model in MODELS:
        model_name = model["name"]
        triple_col = model["triple_col"]
        status_col = model["status_col"]

        valid_count = (
            group[status_col].apply(normalize_status).eq("VALID")
            & group[triple_col].apply(is_real_triple)
        ).sum()

        row[f"{model_name}_valid_count"] = valid_count

    valid_count_rows.append(row)

valid_count_df = pd.DataFrame(valid_count_rows)


# =========================
# SHEET 3 — VALID_PERCENTAGE
# =========================

valid_percentage_rows = []

for sentence_id, group in df.groupby(ID_COL, sort=False):
    sentence = group[SENTENCE_COL].iloc[0]

    row = {
        "id_frase": sentence_id,
        "frase": sentence,
    }

    for model in MODELS:
        model_name = model["name"]
        triple_col = model["triple_col"]
        status_col = model["status_col"]

        total_triples = group[triple_col].apply(is_real_triple).sum()

        valid_triples = (
            group[status_col].apply(normalize_status).eq("VALID")
            & group[triple_col].apply(is_real_triple)
        ).sum()

        if total_triples == 0:
            percentage = ""
        else:
            percentage = valid_triples / total_triples

        row[f"{model_name}_valid_percentage"] = percentage

    valid_percentage_rows.append(row)

valid_percentage_df = pd.DataFrame(valid_percentage_rows)


# =========================
# EXPORT
# =========================

with pd.ExcelWriter(OUTPUT_FILE, engine="openpyxl") as writer:
    valid_marked_df.to_excel(writer, sheet_name="Valid_marked_triples", index=False)
    valid_count_df.to_excel(writer, sheet_name="Valid_count", index=False)
    valid_percentage_df.to_excel(writer, sheet_name="Valid_percentage", index=False)


# =========================
# FORMATTING
# =========================

wb = load_workbook(OUTPUT_FILE)

header_fill = PatternFill("solid", fgColor="D9EAF7")
header_font = Font(bold=True)
thin_border = Border(
    left=Side(style="thin", color="CCCCCC"),
    right=Side(style="thin", color="CCCCCC"),
    top=Side(style="thin", color="CCCCCC"),
    bottom=Side(style="thin", color="CCCCCC"),
)

for ws in wb.worksheets:
    ws.freeze_panes = "A2"

    for cell in ws[1]:
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        cell.border = thin_border

    for row in ws.iter_rows():
        for cell in row:
            cell.alignment = Alignment(vertical="top", wrap_text=True)
            cell.border = thin_border

    for col in ws.columns:
        col_letter = col[0].column_letter
        header = col[0].value

        if header in ["frase"]:
            ws.column_dimensions[col_letter].width = 60
        elif header and "valid_triples" in str(header):
            ws.column_dimensions[col_letter].width = 45
        else:
            ws.column_dimensions[col_letter].width = 18

    for row in ws.iter_rows():
        ws.row_dimensions[row[0].row].height = 35

# Percentuali in formato %
ws = wb["Valid_percentage"]
for row in ws.iter_rows(min_row=2):
    for cell in row[2:]:
        if isinstance(cell.value, (int, float)):
            cell.number_format = "0.00%"

wb.save(OUTPUT_FILE)

print(f"File creato correttamente: {OUTPUT_FILE}")