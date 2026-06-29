from __future__ import annotations

import argparse
from copy import copy
from pathlib import Path
from typing import Any

from openpyxl import load_workbook
from openpyxl.worksheet.datavalidation import DataValidation


BASE_DIR = Path(__file__).resolve().parent
INPUT_DIR = BASE_DIR / "input"
OUTPUT_DIR = BASE_DIR / "output"


def is_blank(value: Any) -> bool:
    return value is None or str(value).strip() == ""


def normalize_header(value: Any) -> str:
    return str(value).strip().lower()


def status_sort_key(column: str) -> int:
    if column == "Status":
        return 0
    if column.startswith("Status."):
        suffix = column.split(".", 1)[1]
        if suffix.isdigit():
            return int(suffix)
    return -1


def next_status_names(existing_headers: list[str], count: int) -> list[str]:
    return ["Status"] * count


def copy_cell_style(source, target) -> None:
    if source.has_style:
        target._style = copy(source._style)
    if source.number_format:
        target.number_format = source.number_format
    if source.alignment:
        target.alignment = copy(source.alignment)
    if source.protection:
        target.protection = copy(source.protection)


def copy_row_style(ws, source_row: int, target_row: int, max_col: int) -> None:
    ws.row_dimensions[target_row].height = ws.row_dimensions[source_row].height
    for col in range(1, max_col + 1):
        copy_cell_style(ws.cell(source_row, col), ws.cell(target_row, col))


def get_headers(ws) -> dict[str, int]:
    return {
        normalize_header(cell.value): cell.column
        for cell in ws[1]
        if not is_blank(cell.value)
    }


def get_id_sentence_columns(ws) -> tuple[int, int] | None:
    headers = get_headers(ws)
    id_col = headers.get("id")
    sentence_col = headers.get("sentence")

    if id_col is None or sentence_col is None:
        return None

    return id_col, sentence_col


def collect_new_model_values(new_file: Path) -> tuple[list[str], dict[int, dict[str, list[Any]]]]:
    wb = load_workbook(new_file, data_only=True)
    ws = wb.active
    headers = [cell.value for cell in ws[1]]

    header_lookup = {
        normalize_header(value): index + 1
        for index, value in enumerate(headers)
        if not is_blank(value)
    }
    id_col = header_lookup.get("id")
    sentence_col = header_lookup.get("sentence")

    if id_col is None or sentence_col is None:
        raise ValueError("The new annotation file must contain ID/id and sentence columns.")

    model_columns = [
        str(value).strip()
        for value in headers
        if not is_blank(value)
        and normalize_header(value) not in {"id", "sentence"}
    ]

    values_by_id: dict[int, dict[str, list[Any]]] = {}
    current_id = None

    for row in range(2, ws.max_row + 1):
        id_value = ws.cell(row, id_col).value
        if not is_blank(id_value):
            current_id = int(id_value)
            values_by_id.setdefault(current_id, {column: [] for column in model_columns})

        if current_id is None:
            continue

        for model_column in model_columns:
            model_col_idx = header_lookup[normalize_header(model_column)]
            value = ws.cell(row, model_col_idx).value
            if not is_blank(value):
                values_by_id[current_id][model_column].append(value)

    return model_columns, values_by_id


def sentence_blocks(ws, id_col: int) -> list[tuple[int, int, int]]:
    starts: list[tuple[int, int]] = []

    for row in range(2, ws.max_row + 1):
        value = ws.cell(row, id_col).value
        if not is_blank(value):
            starts.append((int(value), row))

    blocks = []
    for index, (sentence_id, start_row) in enumerate(starts):
        end_row = starts[index + 1][1] - 1 if index + 1 < len(starts) else ws.max_row
        blocks.append((sentence_id, start_row, end_row))

    return blocks


def limit_to_first_sentences(ws, id_col: int, max_sentences: int) -> None:
    blocks = sentence_blocks(ws, id_col)
    if len(blocks) <= max_sentences:
        return

    first_row_to_delete = blocks[max_sentences][1]
    amount = ws.max_row - first_row_to_delete + 1
    ws.delete_rows(first_row_to_delete, amount)


def list_validation_formula(wb) -> str:
    if "List" not in wb.sheetnames:
        raise ValueError("The workbook must contain a 'List' sheet for status dropdown values.")

    ws = wb["List"]
    last_row = 1
    for row in range(1, ws.max_row + 1):
        if not is_blank(ws.cell(row, 1).value):
            last_row = row

    return f"List!$A$1:$A${last_row}"


def find_template_columns(ws) -> tuple[int, int]:
    result_col = 1
    status_col = 1

    for cell in ws[1]:
        value = str(cell.value).strip() if cell.value is not None else ""
        if status_sort_key(value) >= 0:
            status_col = cell.column
        elif value.lower() not in {"id", "sentence", ""}:
            result_col = cell.column

    return result_col, status_col


def ensure_new_columns(ws, model_columns: list[str]) -> tuple[dict[str, int], list[int]]:
    headers = [cell.value for cell in ws[1]]
    normalized_to_col = {
        normalize_header(cell.value): cell.column
        for cell in ws[1]
        if not is_blank(cell.value)
    }

    result_template_col, status_template_col = find_template_columns(ws)
    status_names = next_status_names([str(value) for value in headers if value is not None], len(model_columns))

    model_col_indexes: dict[str, int] = {}
    status_col_indexes: list[int] = []
    next_col = ws.max_column + 1

    for model_column, status_name in zip(model_columns, status_names):
        existing_model_col = normalized_to_col.get(normalize_header(model_column))
        if existing_model_col is not None:
            model_col = existing_model_col
        else:
            model_col = next_col
            next_col += 1
            ws.cell(1, model_col).value = model_column
            ws.column_dimensions[ws.cell(1, model_col).column_letter].width = (
                ws.column_dimensions[ws.cell(1, result_template_col).column_letter].width
            )

        status_col = next_col
        next_col += 1
        ws.cell(1, status_col).value = status_name
        ws.column_dimensions[ws.cell(1, status_col).column_letter].width = (
            ws.column_dimensions[ws.cell(1, status_template_col).column_letter].width
        )

        for row in range(1, ws.max_row + 1):
            copy_cell_style(ws.cell(row, result_template_col), ws.cell(row, model_col))
            copy_cell_style(ws.cell(row, status_template_col), ws.cell(row, status_col))

        ws.cell(1, model_col).value = model_column
        ws.cell(1, status_col).value = status_name

        model_col_indexes[model_column] = model_col
        status_col_indexes.append(status_col)

    return model_col_indexes, status_col_indexes


def apply_status_dropdown(ws, status_cols: list[int], formula: str) -> None:
    if not status_cols or ws.max_row < 2:
        return

    validation = DataValidation(type="list", formula1=formula, allow_blank=True)
    ws.add_data_validation(validation)

    for col in status_cols:
        column_letter = ws.cell(1, col).column_letter
        validation.add(f"{column_letter}2:{column_letter}{ws.max_row}")


def merge_sheet(
    wb,
    ws,
    model_columns: list[str],
    values_by_id: dict[int, dict[str, list[Any]]],
    max_sentences: int,
) -> None:
    id_sentence = get_id_sentence_columns(ws)
    if id_sentence is None:
        return

    id_col, _sentence_col = id_sentence
    limit_to_first_sentences(ws, id_col, max_sentences)

    model_col_indexes, status_col_indexes = ensure_new_columns(ws, model_columns)
    max_col = ws.max_column

    for sentence_id, start_row, end_row in reversed(sentence_blocks(ws, id_col)):
        new_values = values_by_id.get(sentence_id, {})
        current_len = end_row - start_row + 1
        needed_len = max(
            current_len,
            max((len(new_values.get(model_column, [])) for model_column in model_columns), default=0),
            1,
        )

        if needed_len > current_len:
            insert_at = end_row + 1
            amount = needed_len - current_len
            ws.insert_rows(insert_at, amount)
            for row in range(insert_at, insert_at + amount):
                copy_row_style(ws, end_row, row, max_col)
                for col in range(1, max_col + 1):
                    ws.cell(row, col).value = None

    for sentence_id, start_row, end_row in sentence_blocks(ws, id_col):
        new_values = values_by_id.get(sentence_id, {})

        for row in range(start_row, end_row + 1):
            for model_column in model_columns:
                ws.cell(row, model_col_indexes[model_column]).value = None
            for status_col in status_col_indexes:
                ws.cell(row, status_col).value = None

        for offset, model_column in enumerate(model_columns):
            values = new_values.get(model_column, [])
            model_col = model_col_indexes[model_column]

            for index, value in enumerate(values):
                row = start_row + index
                if row <= end_row:
                    ws.cell(row, model_col).value = value

    apply_status_dropdown(ws, status_col_indexes, list_validation_formula(wb))


def merge_workbooks(old_file: Path, new_file: Path, output_file: Path, max_sentences: int) -> None:
    wb = load_workbook(old_file)
    model_columns, values_by_id = collect_new_model_values(new_file)

    for ws in wb.worksheets:
        merge_sheet(
            wb=wb,
            ws=ws,
            model_columns=model_columns,
            values_by_id=values_by_id,
            max_sentences=max_sentences,
        )

    OUTPUT_DIR.mkdir(exist_ok=True)
    output_file.parent.mkdir(parents=True, exist_ok=True)
    wb.save(output_file)
    print(f"Saved merged annotation workbook to: {output_file}")


def parse_args() -> argparse.Namespace:
    parser = argparse.ArgumentParser(
        description="Merge a complete annotation workbook with a new model annotation file."
    )
    parser.add_argument(
        "--old",
        type=Path,
        default=INPUT_DIR / "annotation_complete.xlsx",
        help="Complete annotation workbook to preserve and extend.",
    )
    parser.add_argument(
        "--new",
        type=Path,
        default=INPUT_DIR / "newannotation.xlsx",
        help="New annotation workbook containing extra model result columns.",
    )
    parser.add_argument(
        "--output",
        type=Path,
        default=OUTPUT_DIR / "annotation_merged.xlsx",
        help="Merged output workbook.",
    )
    parser.add_argument(
        "--max-sentences",
        type=int,
        default=50,
        help="Maximum number of sentence blocks to keep in annotation sheets.",
    )
    return parser.parse_args()


def main() -> None:
    args = parse_args()
    merge_workbooks(args.old, args.new, args.output, args.max_sentences)


if __name__ == "__main__":
    main()
